"""
library_sqlite.py — библиотека файлов в SQLite (метаданные в БД, оригиналы на диске).

API:
  GET  /api/lib/list         — все файлы
  POST /api/lib/add          — добавить файл (upload)
  POST /api/lib/toggle       — вкл/выкл контекст
  DELETE /api/lib/{id}       — удалить
  POST /api/lib/search       — поиск
"""
from __future__ import annotations

import hashlib
import sqlite3
from pathlib import Path

from fastapi import APIRouter, File, Form, UploadFile

DB_PATH = Path("data/library.db")
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parents[4]
UPLOADS_DIR = PROJECT_ROOT / "data" / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

router = APIRouter(prefix="/api/lib", tags=["library-v2"])


def _conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def _ensure_column(conn, table: str, column: str, ddl: str) -> None:
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")


def _init():
    c = _conn()
    c.execute(
        """
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            size INTEGER DEFAULT 0,
            type TEXT DEFAULT 'unknown',
            preview TEXT DEFAULT '',
            use_in_context INTEGER DEFAULT 1,
            source TEXT DEFAULT 'upload',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    _ensure_column(c, "files", "stored_path", "stored_path TEXT DEFAULT ''")
    _ensure_column(c, "files", "sha256", "sha256 TEXT DEFAULT ''")
    c.execute("CREATE INDEX IF NOT EXISTS idx_lib_ctx ON files(use_in_context)")
    c.commit()
    c.close()


_init()


def _safe_disk_name(filename: str, data: bytes) -> str:
    original = Path(filename or "unknown").name
    stem = Path(original).stem or "file"
    suffix = Path(original).suffix
    safe_stem = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in stem)[:80] or "file"
    digest = hashlib.sha256(data).hexdigest()[:12]
    return f"{safe_stem}_{digest}{suffix}"


def _extract_preview(filename: str, contents: bytes) -> str:
    ext = Path(filename).suffix.lower()
    preview = ""
    text_exts = {".txt", ".md", ".json", ".js", ".jsx", ".ts", ".tsx", ".py", ".css", ".html", ".yml", ".yaml", ".xml", ".csv", ".log", ".ini", ".toml", ".bas", ".vbs", ".vba", ".cls", ".frm", ".rsc", ".bat", ".cmd", ".ps1", ".sh", ".sql", ".rb", ".php", ".java", ".c", ".cpp", ".h", ".cs", ".go", ".rs"}
    if ext in text_exts:
        try:
            preview = contents.decode("utf-8", errors="replace")[:12000]
        except Exception:
            preview = ""
    elif ext == ".pdf":
        try:
            import io
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(contents))
            parts = []
            for page in reader.pages[:20]:
                parts.append(page.extract_text() or "")
            preview = "\n".join(parts)[:12000]
        except Exception:
            preview = ""
    elif ext in (".docx", ".doc"):
        try:
            import io
            from docx import Document

            doc = Document(io.BytesIO(contents))
            preview = "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:12000]
        except Exception:
            preview = ""
    elif ext in (".xlsx", ".xls", ".xlsm"):
        try:
            import io
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                parts.append(f"=== {sheet} ===")
                for row in ws.iter_rows(max_row=100, values_only=True):
                    parts.append(" | ".join(str(c) if c is not None else "" for c in row))
            preview = "\n".join(parts)[:12000]
            wb.close()
        except Exception:
            preview = ""
    return preview


@router.get("/list")
def list_files():
    c = _conn()
    rows = c.execute("SELECT * FROM files ORDER BY created_at DESC, id DESC").fetchall()
    c.close()
    return {"ok": True, "items": [dict(r) for r in rows], "count": len(rows)}


@router.post("/add")
async def add_file(
    file: UploadFile = File(...),
    use_in_context: bool = Form(True),
):
    """Загружает файл: оригинал сохраняется на диск, метаданные и preview — в SQLite."""
    contents = await file.read()
    filename = file.filename or "unknown"
    if not contents:
        return {"ok": False, "error": "empty file"}

    disk_name = _safe_disk_name(filename, contents)
    disk_path = UPLOADS_DIR / disk_name
    disk_path.write_bytes(contents)
    preview = _extract_preview(filename, contents)
    sha256 = hashlib.sha256(contents).hexdigest()

    c = _conn()
    cur = c.execute(
        "INSERT INTO files (name, size, type, preview, use_in_context, source, stored_path, sha256) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (
            filename,
            len(contents),
            file.content_type or Path(filename).suffix.lower() or "unknown",
            preview,
            1 if use_in_context else 0,
            "upload",
            str(disk_path),
            sha256,
        ),
    )
    file_id = cur.lastrowid
    c.commit()
    c.close()

    return {
        "ok": True,
        "id": file_id,
        "name": filename,
        "preview_len": len(preview),
        "stored_path": str(disk_path),
    }


@router.post("/toggle")
def toggle_context(file_id: int = Form(...), enabled: bool = Form(True)):
    c = _conn()
    c.execute("UPDATE files SET use_in_context = ? WHERE id = ?", (1 if enabled else 0, file_id))
    c.commit()
    c.close()
    return {"ok": True, "id": file_id, "use_in_context": enabled}


@router.delete("/{file_id}")
def delete_file(file_id: int):
    c = _conn()
    row = c.execute("SELECT stored_path FROM files WHERE id = ?", (file_id,)).fetchone()
    c.execute("DELETE FROM files WHERE id = ?", (file_id,))
    c.commit()
    c.close()

    if row and row["stored_path"]:
        try:
            path = Path(row["stored_path"])
            if path.exists() and path.is_file():
                path.unlink()
        except Exception:
            pass
    return {"ok": True, "deleted_id": file_id}


@router.post("/search")
def search_files(query: str = Form("")):
    q = f"%{query.strip()}%"
    c = _conn()
    rows = c.execute(
        "SELECT * FROM files WHERE name LIKE ? OR preview LIKE ? ORDER BY created_at DESC, id DESC LIMIT 20",
        (q, q),
    ).fetchall()
    c.close()
    return {"ok": True, "items": [dict(r) for r in rows], "count": len(rows)}


@router.get("/context")
def get_context_files():
    """Возвращает файлы с включённым контекстом для LLM."""
    c = _conn()
    rows = c.execute(
        "SELECT id, name, preview, stored_path FROM files WHERE use_in_context = 1 AND preview != '' ORDER BY created_at DESC, id DESC LIMIT 10"
    ).fetchall()
    c.close()
    return {"ok": True, "items": [dict(r) for r in rows], "count": len(rows)}
