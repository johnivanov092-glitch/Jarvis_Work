"""
files.py — извлечение текста из файлов.

Поддержка: PDF, DOCX, XLSX, ZIP, BAS, VBA, CLS, FRM, RSC, и все текстовые.
"""
import io
import zipfile
from pathlib import Path

from fastapi import APIRouter, UploadFile, File
from fastapi.responses import JSONResponse

router = APIRouter(prefix="/api/files", tags=["files"])

# Расширения которые читаем как текст
TEXT_EXTS = {
    ".txt", ".md", ".json", ".js", ".jsx", ".ts", ".tsx", ".py",
    ".css", ".html", ".htm", ".yml", ".yaml", ".xml", ".csv",
    ".log", ".ini", ".toml", ".cfg", ".conf", ".env",
    ".bas", ".vbs", ".vba", ".cls", ".frm", ".rsc",  # VBA / Basic
    ".bat", ".cmd", ".ps1", ".sh",  # Скрипты
    ".sql", ".rb", ".php", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".go", ".rs", ".swift", ".kt", ".r", ".m", ".lua",
    ".pl", ".tcl", ".asm",  # Языки
    ".gitignore", ".dockerfile", ".makefile",
    ".sln", ".csproj", ".pom", ".gradle",  # Проектные
}


def _extract_pdf(data: bytes, max_chars: int = 50000) -> str:
    """Умное извлечение: pypdf → pdfplumber → OCR."""
    try:
        from app.services.pdf_pro import extract_pdf_smart
        result = extract_pdf_smart(data, max_chars)
        text = result.get("text", "")
        # Добавляем таблицы в текст
        tables = result.get("tables", [])
        if tables:
            table_lines = ["\n\n=== ТАБЛИЦЫ ==="]
            for t in tables[:5]:
                headers = t.get("headers", [])
                rows = t.get("rows", [])
                table_lines.append(f"\nТаблица (стр. {t.get('page', '?')}):")
                if headers:
                    table_lines.append(" | ".join(str(h or "") for h in headers))
                    table_lines.append("-" * 40)
                for row in rows[:20]:
                    table_lines.append(" | ".join(str(c or "") for c in row))
            text += "\n".join(table_lines)
        method = result.get("method", "")
        if result.get("ocr_used"):
            text = f"[OCR распознавание]\n{text}"
        return text[:max_chars]
    except ImportError:
        # Fallback на простой pypdf
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            parts, total = [], 0
            for page in reader.pages:
                t = page.extract_text() or ""
                if total + len(t) > max_chars:
                    parts.append(t[:max_chars - total])
                    break
                parts.append(t)
                total += len(t)
            return "\n\n".join(parts)
        except ImportError:
            return "[pypdf не установлен: pip install pypdf]"
    except Exception as e:
        return f"[PDF ошибка: {e}]"


def _extract_docx(data: bytes, max_chars: int = 30000) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        parts = []
        total = 0
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                if total + len(text) > max_chars:
                    break
                parts.append(text)
                total += len(text)
        return "\n".join(parts)
    except ImportError:
        return "[python-docx не установлен: pip install python-docx]"
    except Exception as e:
        return f"[DOCX ошибка: {e}]"


def _extract_xlsx(data: bytes, max_chars: int = 30000) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        total = 0
        for sheet in wb.sheetnames[:5]:  # Макс 5 листов
            ws = wb[sheet]
            parts.append(f"=== Лист: {sheet} ===")
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells)
                if total + len(line) > max_chars:
                    break
                parts.append(line)
                total += len(line)
        wb.close()
        return "\n".join(parts)
    except ImportError:
        return "[openpyxl не установлен: pip install openpyxl]"
    except Exception as e:
        return f"[XLSX ошибка: {e}]"


def _extract_zip(data: bytes, max_chars: int = 30000) -> str:
    """Открывает ZIP и читает текстовые файлы внутри."""
    try:
        parts = []
        total = 0
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            parts.append(f"ZIP содержит {len(zf.namelist())} файлов:")
            for name in zf.namelist()[:30]:  # Макс 30 файлов
                ext = Path(name).suffix.lower()
                size = zf.getinfo(name).file_size
                parts.append(f"  - {name} ({size} байт)")

                # Читаем текстовые файлы
                if ext in TEXT_EXTS and size < 100_000:
                    try:
                        content = zf.read(name).decode("utf-8", errors="replace")
                        if total + len(content) > max_chars:
                            content = content[:max_chars - total]
                        parts.append(f"\n--- {name} ---\n{content}")
                        total += len(content)
                    except Exception:
                        pass

                if total > max_chars:
                    break

        return "\n".join(parts)
    except Exception as e:
        return f"[ZIP ошибка: {e}]"


def _extract_text(data: bytes, max_chars: int = 30000) -> str:
    """Читает текст, пробуя несколько кодировок (UTF-8 → CP1251 → CP866)."""
    if not data:
        return ""
    for enc in ("utf-8", "utf-8-sig", "cp1251", "cp866", "latin-1"):
        try:
            text = data.decode(enc)
            # Проверяем что текст читаемый (нет замен)
            if "\ufffd" not in text[:500]:
                return text[:max_chars]
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace")[:max_chars]


@router.post("/extract-text")
async def extract_text(file: UploadFile = File(...)):
    """Извлекает текст из любого поддерживаемого файла."""
    try:
        contents = await file.read()
        filename = (file.filename or "").strip()
        ext = Path(filename).suffix.lower()

        if ext == ".pdf":
            text = _extract_pdf(contents)
        elif ext in (".docx", ".doc"):
            text = _extract_docx(contents)
        elif ext in (".xlsx", ".xls", ".xlsm"):
            text = _extract_xlsx(contents)
        elif ext == ".zip":
            text = _extract_zip(contents)
        else:
            text = _extract_text(contents)

        return {
            "ok": True,
            "filename": filename,
            "size": len(contents),
            "text": text,
            "chars": len(text),
            "type": ext,
        }
    except Exception as exc:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(exc)})
